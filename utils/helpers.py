import pandas as pd
import streamlit as st
import re
from datetime import datetime
import requests
from utils.metrics import *
import pandas as pd
from openpyxl.styles import PatternFill
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

def extract_update_date(file_url):
    """Extract and format the last update date from filename in content-disposition header."""
    response = requests.get(file_url)
    content_disposition = response.headers.get('content-disposition', '')
    filename = re.findall("filename=(.+)", content_disposition)
    filename = filename[0] if filename else "Unknown"
    
    match = re.search(r"(\d{2})_(\d{2})_(\d{4})", filename)
    if match:
        day, month, year = match.groups()
        date_obj = datetime.strptime(f"{day}-{month}-{year}", "%d-%m-%Y")
        return date_obj.strftime("%B %d, %Y")
    return "Unknown"

def load_club_performance_data(secret_key: str) -> pd.DataFrame:
    """
    Loads Club Performance data from Google Drive using a secret key.
    
    Args:
        secret_key (str): The key to access the file ID from Streamlit secrets.

    Returns:
        pd.DataFrame: Cleaned DataFrame with valid club names.
    """
    try:
        file_id = st.secrets[secret_key]
        gsheet_url = f"https://drive.google.com/uc?export=download&id={file_id}"

        df = pd.read_csv(gsheet_url)
        try:
            update_date = df.iloc[-1]['Division'][-10:]
        except Exception:
            update_date = "Not available"

        df = df[df["Club Name"].notna()]  # Filter rows with non-empty club names
        df["Club Number"] = df["Club Number"].astype(int)
        return df, update_date

    except Exception as e:
        st.warning(f"Could not load Club Performance data: {e}")
        return pd.DataFrame(), 'January 01, 1900'  # Return empty DataFrame on failure

def get_quarter_delta(df_latest: pd.DataFrame, 
                      df_last_quarter: pd.DataFrame, 
                      cols_to_diff: list[str],
                      merge_on: str = "Club Number") -> pd.DataFrame:
    """
    Compute quarter-only data by subtracting last quarter snapshot from the latest YTD data.

    Parameters:
        df_latest (pd.DataFrame): Latest YTD data (e.g., Q2)
        df_last_quarter (pd.DataFrame): Snapshot as of end of last quarter (e.g., Q1)
        cols_to_diff (list[str]): List of metric columns to compute delta on
        merge_on (str): Column to merge on (default = "Club Number")

    Returns:
        pd.DataFrame: New dataframe with quarter-only values
    """

    # Standardize merge key (e.g., strip whitespace, uppercase)
    df_latest[merge_on] = df_latest[merge_on].astype(str).str.strip().str.upper()
    df_last_quarter[merge_on] = df_last_quarter[merge_on].astype(str).str.strip().str.upper()

    # Merge both snapshots on Club Number
    df_merged = df_latest.merge(
        df_last_quarter,
        on=merge_on,
        suffixes=("_latest", "_q1"),
        how="left"
    )

    # Subtract old values from latest
    for col in cols_to_diff:
        col_latest = f"{col}_latest"
        col_q1 = f"{col}_q1"
        df_merged[col] = df_merged.get(col_latest, 0) - df_merged.get(col_q1, 0).fillna(0)

    # Keep only Club Number and computed delta columns
    return df_merged[[merge_on] + cols_to_diff]

def get_csp_improvement(df_latest: pd.DataFrame, df_last_quarter: pd.DataFrame) -> pd.DataFrame:
    """
    Returns a DataFrame with Club Number and updated CSP value:
    - 'Y' if current CSP is 'Y' and previous was 'N' or missing.
    - 'N' otherwise.
    """

    # Extract and rename columns from last quarter
    prev_csp = df_last_quarter[['Club Number', 'CSP']].rename(columns={'CSP': 'CSP_Previous'})
    
    # Merge current with previous
    merged = df_latest[['Club Number', 'CSP']].merge(prev_csp, on='Club Number', how='left')

    # Apply logic to determine improvement
    merged['CSP'] = merged.apply(
        lambda row: 'Y' if row['CSP'] == 'Y' and row.get('CSP_Previous', 'N') != 'Y' else 'N',
        axis=1
    )
    merged['Club Number'] = merged['Club Number'].astype(int)
    return merged[['Club Number', 'CSP']]

# ------------------ Load and Prepare Data ------------------ #
def load_data_club_performance(gsheet_url=None):

    cq = st.secrets["Current_Quarter"]

    # Map current to last quarter
    quarter_map = {
        "Q1": None,
        "Q2": "Q1",
        "Q3": "Q2",
        "Q4": "Q3"
    }
    lq = quarter_map.get(cq)

    # Load common data
    df_base, quarter_base_date = load_club_performance_data(secret_key="GOOGLE_DRIVE_FILE_ID_CLUB_PERFORMANCE_BASE_" + cq)
    df_latest, update_date = load_club_performance_data(secret_key="GOOGLE_DRIVE_FILE_ID_CLUB_PERFORMANCE_" + cq)

    # Column groups
    base_col = ['District', 'Division', 'Area', 'Club Number', 'Club Name',
                'Club Status', 'Mem. Base', 'Active Members', 'Net Growth']

    other_col = ['Club Number', 'CSP', 'Goals Met', 'Level 1s', 'Level 2s', 'Add. Level 2s', 'Level 3s',
                'Level 4s, Path Completions, or DTM Awards',
                'Add. Level 4s, Path Completions, or DTM award', 'New Members',
                'Add. New Members', 'Off. Trained Round 1', 'Off. Trained Round 2',
                'Mem. dues on time Oct', 'Mem. dues on time Apr', 'Off. List On Time',
                'Club Distinguished Status']

    # Compute current quarter-only data
    if lq is None:
        # First quarter — use latest as-is
        df_current_only = df_latest[other_col].copy()
    else:
        # Load last quarter data and compute delta
        df_last_quarter, _ = load_club_performance_data(secret_key="GOOGLE_DRIVE_FILE_ID_CLUB_PERFORMANCE_" + lq)
        df_current_only = get_quarter_delta(
            df_latest=df_latest,
            df_last_quarter=df_last_quarter,
            cols_to_diff=[x for x in other_col if x not in ["Club Number", "CSP"]],
            merge_on="Club Number"
        )
        df_current_only["Club Number"] = df_current_only["Club Number"].astype(int)

    df = df_base[base_col].merge(df_current_only, on='Club Number', how='left')

    df_updated_csp = get_csp_improvement(df_latest, df_last_quarter)
    df = df.merge(df_updated_csp, on='Club Number', how='left')

    new_clubs = df_current_only[~df_current_only["Club Number"].isin(df_base["Club Number"])]

    df = pd.concat([df, new_clubs], ignore_index=True)

    df["Club Number"] = df["Club Number"].astype(int)

    df_edu_achievements = load_excel_data("GOOGLE_DRIVE_FILE_ID_EDU_ACHIEVEMENTS", ["Club", "Name", "Award", "Date"], sheet_name="Sheet1")
    df_edu_achievements.rename(columns={"Club": "Club Number"}, inplace=True)
    df_tc = load_excel_data("GOOGLE_DRIVE_FILE_ID_TRIPLE_CROWN", ["Club Name", "Member"], sheet_name="300925")
    df_tc = df[['Club Name', 'Club Number']].merge(df_tc, how = 'inner')
    df = calculate_points(df, df_edu_achievements, df_tc)
    df = assign_grouping(df)
    # df = df[df['Group'] != 'Unknown']
    return df, update_date

def load_csv_from_secret(secret_key: str, columns: list[str]) -> pd.DataFrame:
    """
    Loads a CSV from Google Drive using a file ID stored in Streamlit secrets.
    If loading fails, returns an empty DataFrame with the given columns.
    """
    try:
        file_id = st.secrets[secret_key]
        gsheet_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv"
        df = pd.read_csv(gsheet_url)
    except Exception as e:
        # st.warning(f"Could not load file for {secret_key}: {e}")
        df = pd.DataFrame(columns=columns)
    return df

def load_excel_data(secret_key: str, columns: list[str], sheet_name="Sheet1") -> pd.DataFrame:
    """
    Loads a CSV from Google Drive using a file ID stored in Streamlit secrets.
    If loading fails, returns an empty DataFrame with the given columns.
    """
    try:
        file_id = st.secrets[secret_key]
        gsheet_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        df = pd.read_excel(gsheet_url, sheet_name=sheet_name, skiprows=1)
    except Exception as e:
        # st.warning(f"Could not load Education Achievements data: {e}")
        df = pd.DataFrame(columns=columns)
    return df

def prepare_pathways_pioneers_data(df_club_performance):
    """
    Process club performance data and merge with contest data to create pathways pioneers leaderboard.
    
    Args:
        df_club_performance: DataFrame with club performance data
        
    Returns:
        DataFrame with processed pathways pioneers data
    """
    df = df_club_performance.copy()

    df_contests = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_CONTESTS", ["Select Your Club", "Humorous Contest", "TableTopics Contest", "Evaluation Contest", "International Contest"])
    
    contests_points = calculate_contest_points(df_contests)

    df_pathways_pioneers = df.merge(contests_points, left_on="Club Number", right_on="Club Number", how="left")

    df_pathways_pioneers = df_pathways_pioneers.fillna(0)

    # Add tier points
    df_pathways_pioneers['Pathways Pioneers'] = (
        df_pathways_pioneers[['L1 Points', 'L2 Points', 'L3 Points', 'L4 Points', 'L5 Points', 'DTM Points', 'TC Points', 
               'Humorous Contest', 'TableTopics Contest', 'Evaluation Contest', 'International Contest']].sum(axis=1)
    )

    # Select columns and format
    columns = ['Club Name', 'Club Number', 'Club Group', 'Active Members', 'Pathways Pioneers',
               'L1 Points', 'L2 Points', 'L3 Points', 'L4 Points', 'L5 Points', 'DTM Points', 'TC Points',
               'Humorous Contest', 'TableTopics Contest', 'Evaluation Contest', 'International Contest']
    
    # Sort and reset index
    df_pathways_pioneers = df_pathways_pioneers[df_pathways_pioneers['Active Members'] >= 8]
    return df_pathways_pioneers[columns].sort_values(by='Club Name').reset_index(drop=True)

def prepare_leadership_innovators_data(df_club_performance):
    """
    Process club performance data and merge with MOT data to create leadership innovators leaderboard.
    
    Args:
        df_club_performance: DataFrame with club performance data
        
    Returns:
        DataFrame with processed leadership innovators data
    """
    df = df_club_performance.copy()

    # Load and process MOT data
    df_mot = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_MOMENTS_OF_TRUTH", ["Select Your Club", "MOT_Q1", "MOT_Q3"])
    df_mot_scores = mot_scores(df_mot)

    df_leadership_innovators = df.merge(df_mot_scores, left_on="Club Number", right_on="Club Number", how="left")

    # Load and process PCC data
    df_pcc = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_PATHWAYS_COMPLETION_CELEBRATION", ["Select Your Club", "Pathways_Completion_Celebration"])
    df_pcc_scores = pathways_completion_scores(df_pcc)
    df_leadership_innovators = df_leadership_innovators.merge(df_pcc_scores, left_on="Club Number", right_on="Club Number", how="left")

    # Load and process Mentorship Program data
    df_mp = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_MENTORSHIP_PROGRAM", ["Select Your Club", "Mentorship_Programme"])
    df_mp_scores = mentorship_programme_scores(df_mp)
    df_leadership_innovators = df_leadership_innovators.merge(df_mp_scores, left_on="Club Number", right_on="Club Number", how="left")

    # Load and process dcp data
    df_dcp = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_DCP", ["Select Your Club", "Distinguished_Club_Partners"])
    df_dcp_scores = distinguished_club_partners_scores(df_dcp)
    df_leadership_innovators = df_leadership_innovators.merge(df_dcp_scores, left_on="Club Number", right_on="Club Number", how="left")

    # Load and process sth data
    df_sth = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_STH", ["Select Your Club", "Successful_Transition_Handover"])
    df_sth_scores = successful_handover_scores(df_sth)
    df_leadership_innovators = df_leadership_innovators.merge(df_sth_scores, left_on="Club Number", right_on="Club Number", how="left")

    df_leadership_innovators['President_Distinguished'] = 0
    df_leadership_innovators['Smedley_Distinguished'] = 0

    df_leadership_innovators = df_leadership_innovators.fillna(0)

    # Add tier points
    df_leadership_innovators['Leadership Innovators'] = (
        df_leadership_innovators[['COT R1 Points', 'COT R2 Points', 'MOT_Q1', 'MOT_Q3', 
                                  'Pathways_Completion_Celebration','Mentorship_Programme',
                                  'President_Distinguished', 'Smedley_Distinguished',
                                  'Distinguished_Club_Partners', 'Successful_Transition_Handover']].sum(axis=1)
    )

    # Select columns and format
    columns = ['Club Name', 'Club Number', 'Club Group', 'Active Members', 'Leadership Innovators',
               'COT R1 Points', 'COT R2 Points', 'MOT_Q1', 'MOT_Q3', 
                'Pathways_Completion_Celebration','Mentorship_Programme',
                'President_Distinguished', 'Smedley_Distinguished',
                'Distinguished_Club_Partners', 'Successful_Transition_Handover']
    
    # Sort and reset index
    df_leadership_innovators = df_leadership_innovators[df_leadership_innovators['Active Members'] >= 8]
    return df_leadership_innovators[columns].sort_values(by='Club Name').reset_index(drop=True)

def prepare_excellence_champions_data(df_club_performance):
    """
    Process club performance data and merge with Club Success Plan data to create excellence champions leaderboard.
    
    Args:
        df_club_performance: DataFrame with club performance data
        
    Returns:
        DataFrame with processed excellence champions data
    """
    df = df_club_performance.copy()

    df_qis = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_QIS", ["Select Your Club", "Quality_Initiatives"])

    df_qis_scores = quality_initiatives_scores(df_qis)
    df_excellence_champions = df.merge(df_qis_scores, left_on="Club Number", right_on="Club Number", how="left")

    df_mo = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_MEMBER_ONBOARDING", ["Select Your Club", "Member_Onboarding"])

    df_mo_scores = member_onboarding_scores(df_mo)
    df_excellence_champions = df_excellence_champions.merge(df_mo_scores, left_on="Club Number", right_on="Club Number", how="left")

    df_excellence_champions["Club_Success_Plan"] = df_excellence_champions["CSP"].apply(
    lambda x: 20 if str(x).strip().upper() == "Y" else 0
    )

    df_excellence_champions['FirstTime_Distinguished'] = 0

    df_pr = load_csv_from_secret("GOOGLE_DRIVE_FILE_ID_MEMBERSHIP_LIST", ["Club Number", "100%_Pathway_Registration"])
    df_pr = pathway_enrollment_scores(df_pr)
    df_excellence_champions = df_excellence_champions.merge(df_pr, left_on="Club Number", right_on="Club Number", how="left")

    # Replace NaN values with 0
    df_excellence_champions = df_excellence_champions.fillna(0)

    # Add tier points
    df_excellence_champions['Excellence Champions'] = (
        df_excellence_champions[['Club_Success_Plan', 'FirstTime_Distinguished', 'Early10_Distinguished', 'Quality_Initiatives', '100%_Pathway_Registration', 'Member_Onboarding']].sum(axis=1)
    )
    
    # Select columns and format
    columns = ['Club Name', 'Club Number', 'Club Group', 'Active Members', 'Excellence Champions', 'Club_Success_Plan', 'FirstTime_Distinguished', 'Early10_Distinguished', 'Quality_Initiatives', '100%_Pathway_Registration', 'Member_Onboarding']

    # Sort and reset index
    df_excellence_champions = df_excellence_champions[df_excellence_champions['Active Members'] >= 8]
    return df_excellence_champions[columns].sort_values(by='Club Name').reset_index(drop=True)

def generate_leaderboard_excel(df_merged: pd.DataFrame, group_meta: dict, incentives_tiers: dict) -> BytesIO:
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for group_key, group_info in group_meta.items():
        group_name = group_info['Name']
        df_group = df_merged[df_merged['Club Group'] == group_name].copy()

        for tier_key, tier_info in incentives_tiers.items():
            tier_name = tier_info['Name']
            sheet_name = f"{group_name[:15]} - {tier_name[:15]}"
            ws = wb.create_sheet(title=sheet_name)

            # Avoid filtering out clubs with 0 points, keep all clubs
            df_sorted = df_group.sort_values(
                by=[tier_name, 'Total Club Points', 'Club Name'],
                ascending=[False, False, True],
                kind="mergesort"
            ).reset_index(drop=True)

            # Add group rank only for clubs with >0 points
            df_sorted["Group Rank"] = None
            active_clubs = df_sorted[tier_name] > 0
            df_sorted.loc[active_clubs, "Group Rank"] = range(1, active_clubs.sum() + 1)

            # Top 3 logic (used only for highlighting)
            highlight_mask = [False] * len(df_sorted)
            df_positive = df_sorted[active_clubs]
            if len(df_positive) >= 3:
                third_score = df_positive.iloc[2][tier_name]
                third_points = df_positive.iloc[2]["Total Club Points"]
                highlight_mask = (
                    (df_sorted[tier_name] > third_score) |
                    ((df_sorted[tier_name] == third_score) &
                     (df_sorted["Total Club Points"] >= third_points))
                ).tolist()
            else:
                highlight_mask = active_clubs.tolist()

            # Final export columns (no Top 3 column)
            df_export = df_sorted[['Club Name', 'Club Group', tier_name, 'Total Club Points']].copy()
            df_export.columns = ['Club Name', 'Club Group', 'Tier Points', 'Total Club Points']

            # Write to Excel
            for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    # Highlight only active clubs marked for top 3
                    if r_idx > 1 and highlight_mask[r_idx - 2]:  # row index adjusted
                        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    wb.save(output)
    output.seek(0)
    return output



